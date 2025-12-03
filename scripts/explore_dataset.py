"""
Explore the dataset to understand its structure and classification.
"""
import openpyxl
from pathlib import Path

excel_path = Path("data/01_raw/datasets/PBCs_microorgansim_information.xlsx")

print("=" * 80)
print("DATASET EXPLORATION")
print("=" * 80)

# Load Excel file
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

print(f"\nExcel file: {excel_path}")
print(f"Sheet name: {sheet.title}")
print(f"Total rows: {sheet.max_row}")
print(f"Total columns: {sheet.max_column}")

print("\n" + "=" * 80)
print("FIRST 30 ROWS:")
print("=" * 80)

for i, row in enumerate(sheet.iter_rows(values_only=True)):
    print(f"Row {i}: {row}")
    if i > 30:
        break
